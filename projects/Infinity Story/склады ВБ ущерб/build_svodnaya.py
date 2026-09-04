import openpyxl, glob, re, os, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

SRC="src/склады ВБ_ущерб Infinity Story"
def norm_size(s): return re.sub(r'\s+','',str(s)).upper().replace('ONESIZE','ONE SIZE')
SIZE_ORDER=['XS','S','M','L','XL','XXL','XS-S','XS-M','M-L','L-XXL','XL-XXL','ONE SIZE']

# ---- load warehouses, ordered by date from filename
files=[]
for f in glob.glob(f"{SRC}/товары*.xlsx"):
    m=re.match(r'товары (\d+)\.(\d+) (.+)\.xlsx', os.path.basename(f))
    d,mo,city=int(m.group(1)),int(m.group(2)),m.group(3)
    city={'ШУШАРЫ':'Шушары','Невиномысск':'Невинномысск','Новосемейкино Самара':'Самара Новосемейкино'}.get(city,city)
    files.append(((mo,d),f"{d:02d}.{mo:02d}",city,f))
files.sort()
whs=[c for _,_,c,_ in files]
wh_dates={c:dt for _,dt,c,_ in files}

rows={}   # key (wb, size) -> dict
src_totals={}
for _,dt,city,f in files:
    ws=openpyxl.load_workbook(f,read_only=True,data_only=True).worksheets[0]
    data=list(ws.iter_rows(values_only=True))
    idx={h:i for i,h in enumerate(data[0]) if h}; whcol=idx[[h for h in data[0] if h][-1]]
    tot=0
    for r in data[1:]:
        if r[idx['Артикул WB']] is None: continue
        wb=str(r[idx['Артикул WB']]).strip(); size=norm_size(r[idx['Размер вещи']])
        k=(wb,size); d=rows.setdefault(k,dict(wb=wb,size=size,brand=None,subj=None,art=None,bc=None,qty=collections.Counter()))
        d['art']=d['art'] or r[idx['Артикул продавца']].strip()
        if 'Бренд' in idx and r[idx['Бренд']]: d['brand']=d['brand'] or r[idx['Бренд']]
        if 'Предмет' in idx and r[idx['Предмет']]: d['subj']=d['subj'] or r[idx['Предмет']]
        if 'Баркод' in idx and r[idx['Баркод']]: d['bc']=d['bc'] or str(r[idx['Баркод']]).strip()
        q=int(r[whcol]); d['qty'][city]+=q; tot+=q
    src_totals[city]=tot

# fill subject/brand by article from other rows; fall back to price file subject
pw=openpyxl.load_workbook(f"{SRC}/Цены ВБ.xlsx",read_only=True).worksheets[0]
prices={}; psubj={}
for r in list(pw.iter_rows(values_only=True))[1:]:
    if r[1]: prices[str(r[1]).strip().lower()]=r[2]; psubj[str(r[1]).strip().lower()]=r[0]
by_art=collections.defaultdict(list)
for d in rows.values(): by_art[d['art'].lower()].append(d)
for a,ds in by_art.items():
    subj=next((d['subj'] for d in ds if d['subj']),None) or psubj.get(a)
    for d in ds: d['subj']=d['subj'] or subj; d['brand']=d['brand'] or 'Infinity Story'

def size_key(s): return SIZE_ORDER.index(s) if s in SIZE_ORDER else 99
ordered=sorted(rows.values(), key=lambda d:(d['subj'] or '', d['art'].lower(), size_key(d['size'])))

# ---- workbook
wb=Workbook(); ws=wb.active; ws.title="Сводная"
F=lambda **k: Font(name='Arial',size=10,**k)
thin=Side(style='thin',color='BFBFBF'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
hdr_fill=PatternFill('solid',fgColor='1F3864'); wh_fill=PatternFill('solid',fgColor='2E75B6')
tot_fill=PatternFill('solid',fgColor='D9E1F2'); yellow=PatternFill('solid',fgColor='FFFF00'); grey=PatternFill('solid',fgColor='F2F2F2')

fixed=['Бренд','Предмет','Артикул','Артикул WB','Баркод','Размер']
headers=fixed+whs+['Цена','Итого шт','Итого ₽']
ws.append(headers)
nf=len(fixed); first_wh=nf+1; last_wh=nf+len(whs)
c_price=last_wh+1; c_qty=last_wh+2; c_sum=last_wh+3
for i,h in enumerate(headers,1):
    c=ws.cell(1,i); c.font=Font(name='Arial',size=10,bold=True,color='FFFFFF'); c.border=border
    c.fill=wh_fill if first_wh<=i<=last_wh else hdr_fill
    c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
for i,city in enumerate(whs,first_wh):
    ws.cell(1,i).comment=Comment(f"Файл: товары {wh_dates[city]} … (дата ущерба {wh_dates[city]})","Сводная")
ws.row_dimensions[1].height=45

no_price=[]; no_bc=[]
r=2
for d in ordered:
    vals=[d['brand'],d['subj'],d['art'],d['wb'],d['bc'] or '',d['size']]+[d['qty'].get(c) or None for c in whs]
    for i,v in enumerate(vals,1):
        c=ws.cell(r,i,v); c.font=F(); c.border=border
        if i>=first_wh: c.alignment=Alignment(horizontal='center')
        if i in (4,5): c.number_format='@'
    if not d['bc']: ws.cell(r,5).fill=yellow; no_bc.append(d)
    p=prices.get(d['art'].lower())
    pc=ws.cell(r,c_price,p); pc.font=F(color='0000FF'); pc.border=border; pc.number_format='#,##0'
    if p is None: pc.fill=yellow; no_price.append(d)
    a,b=get_column_letter(first_wh),get_column_letter(last_wh); P,Q=get_column_letter(c_price),get_column_letter(c_qty)
    qc=ws.cell(r,c_qty,f"=SUM({a}{r}:{b}{r})"); qc.font=F(bold=True); qc.border=border; qc.alignment=Alignment(horizontal='center')
    sc=ws.cell(r,c_sum,f'=IF({P}{r}="","",{P}{r}*{Q}{r})'); sc.font=F(bold=True); sc.border=border; sc.number_format='#,##0'
    r+=1
last=r-1
# totals row
ws.cell(r,1,"ИТОГО").font=F(bold=True)
for i in range(1,c_sum+1):
    c=ws.cell(r,i); c.fill=tot_fill; c.border=border; c.font=F(bold=True)
    if i>=first_wh:
        L=get_column_letter(i); c.value=f"=SUM({L}2:{L}{last})"; c.alignment=Alignment(horizontal='center')
        if i in (c_sum,c_price): c.number_format='#,##0'
ws.cell(r,c_price).value=None
tot_row=r
# legend
r+=2
Q=get_column_letter(c_qty); P=get_column_letter(c_price)
legend=[("Как читать таблицу",True),
 ("Столбцы складов = кол-во повреждённого товара по отчёту ВБ; название склада взято из имени файла, дата акта — в примечании к заголовку столбца.",False),
 ("Цена — из файла «Цены ВБ» (синий текст = можно править). Итого ₽ = Цена × Итого шт. Итоги пересчитываются автоматически.",False),
 (f"Жёлтая ячейка цены — артикула нет в «Цены ВБ»: впиши цену, сумма посчитается сама. Таких строк: {len(no_price)}, шт без цены: ",False),
 (f"Жёлтая ячейка баркода — в отчётах Краснодар / Невинномысск / Шушары баркод не выгружался, и такой пары «артикул WB + размер» нет на других складах. Строк: {len(no_bc)}.",False),
 ("Размеры «ONESIZE» и «ONE SIZE» объединены в «ONE SIZE». Предмет для КОФТАДИОР взят из выгрузки ВБ («Кофты»), в прайсе он числится как «Кардиганы».",False)]
for txt,b in legend:
    c=ws.cell(r,1,txt); c.font=F(bold=b,italic=not b,color='595959' if not b else None)
    r+=1
# qty without price formula
nq=ws.cell(r-3,c_qty,f'=SUMIF({P}2:{P}{last},"",{Q}2:{Q}{last})'); nq.font=F(bold=True,color='C00000')

widths={1:14,2:12,3:30,4:12,5:15,6:10}
for i in range(1,c_sum+1): ws.column_dimensions[get_column_letter(i)].width=widths.get(i,11 if i<=last_wh else 12)
ws.column_dimensions[get_column_letter(c_sum)].width=13
ws.freeze_panes=ws.cell(2,first_wh)
ws.auto_filter.ref=f"A1:{get_column_letter(c_sum)}{last}"

# ---- check sheet
cs=wb.create_sheet("Сверка")
cs.append(["Склад","Дата акта","Итого в исходном файле","Итого в сводной","Разница"])
for i in range(1,6):
    c=cs.cell(1,i); c.font=Font(name='Arial',size=10,bold=True,color='FFFFFF'); c.fill=hdr_fill; c.border=border; c.alignment=Alignment(horizontal='center',wrap_text=True)
for j,city in enumerate(whs):
    rr=j+2; L=get_column_letter(first_wh+j)
    cs.cell(rr,1,city); cs.cell(rr,2,wh_dates[city]); cs.cell(rr,3,src_totals[city]).font=F(color='0000FF')
    cs.cell(rr,4,f"=Сводная!{L}{tot_row}"); cs.cell(rr,5,f"=D{rr}-C{rr}")
    for i in range(1,6):
        c=cs.cell(rr,i); c.border=border
        if i!=3: c.font=F()
        if i>1: c.alignment=Alignment(horizontal='center')
rr=len(whs)+2
cs.cell(rr,1,"ИТОГО"); 
for i,fml in ((3,f"=SUM(C2:C{rr-1})"),(4,f"=SUM(D2:D{rr-1})"),(5,f"=D{rr}-C{rr}")): cs.cell(rr,i,fml)
for i in range(1,6):
    c=cs.cell(rr,i); c.font=F(bold=True); c.fill=tot_fill; c.border=border
    if i>1: c.alignment=Alignment(horizontal='center')
cs.cell(rr+2,1,"«Итого в исходном файле» — контрольная строка внизу каждого файла ВБ (введено вручную из файлов). Разница должна быть 0.").font=F(italic=True,color='595959')
for i,w in zip(range(1,6),(24,12,22,18,12)): cs.column_dimensions[get_column_letter(i)].width=w

out="Склады ВБ ущерб Infinity Story — сводная.xlsx"
wb.save(out); print("saved",out,"rows",len(ordered),"warehouses",len(whs))
print("no price:",len(no_price),"no bc:",len(no_bc))
